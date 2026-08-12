from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlsxwriter


BANK_ACCOUNT_HEADER = "BANK AND ACCOUNT NUMBER"
BANK_ACCOUNT_KEY = "__BANK_ACCOUNT__"

DEFAULT_INFLOW_HEADERS = ["DATE", "DESCRIPTION", "DEBIT", "CREDIT", "CLASSIFICATION"]
DEFAULT_OUTFLOW_HEADERS = ["DATE", "DESCRIPTION", "DEBIT", "CONFIRM", "DIFF", "CLASSIFICATION"]

PARSER_BANK_NAMES = {
    "zenith-style": "Zenith Bank",
    "zenith": "Zenith Bank",
    "uba": "UBA",
    "fcmb": "FCMB",
    "providus": "Providus Bank",
    "firstbank": "FirstBank",
    "fidelity": "Fidelity Bank",
    "gtbank": "GTBank",
    "globus": "Globus Bank",
    "lotus": "Lotus Bank",
    "standard-chartered": "Standard Chartered Bank",
    "taj": "TAJ Bank",
    "jaiz": "Jaiz Bank",
    "clear-junction": "Clear Junction",
    "wema-treasure": "Wema Bank",
    "wema": "Wema Bank",
    "moniepoint": "Moniepoint",
    "opay": "OPay",
    "keystone": "Keystone Bank",
}


@dataclass(slots=True)
class WorkbookPreview:
    filename: str
    path: Path
    bank_name: str | None
    account_number: str | None
    parser_name: str | None
    inflow_count: int = 0
    outflow_count: int = 0

    @property
    def needs_manual_details(self) -> bool:
        return not self.bank_name or not self.account_number

    @property
    def bank_account_label(self) -> str:
        parts = [part for part in (self.bank_name, self.account_number) if part]
        return " ".join(parts)


@dataclass(slots=True)
class SheetData:
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclass(slots=True)
class AnalyzedWorkbook:
    preview: WorkbookPreview
    inflows: SheetData
    outflows: SheetData


@dataclass(slots=True)
class ConsolidationResult:
    output_path: Path
    file_count: int
    inflow_count: int
    outflow_count: int
    files: list[WorkbookPreview]


def inspect_analyzed_workbook(path: Path, filename: str | None = None) -> WorkbookPreview:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        bank_name, account_number, parser_name = extract_metadata(workbook)
        inflows = read_transaction_sheet(workbook, "Inflows")
        outflows = read_transaction_sheet(workbook, "Outflows")
        if not inflows.headers and not outflows.headers:
            raise ValueError("Workbook is missing the Inflows and Outflows sheets.")
        inflow_count = len(inflows.rows)
        outflow_count = len(outflows.rows)
    finally:
        workbook.close()

    display_filename = filename or path.name
    bank_name = bank_name or bank_name_from_filename(display_filename)
    return WorkbookPreview(
        filename=display_filename,
        path=path,
        bank_name=bank_name,
        account_number=account_number,
        parser_name=parser_name,
        inflow_count=inflow_count,
        outflow_count=outflow_count,
    )


def consolidate_analyzed_workbooks(
    files: list[WorkbookPreview],
    output_path: Path,
    *,
    detail_overrides: dict[int, tuple[str, str]] | None = None,
) -> ConsolidationResult:
    analyzed_workbooks: list[AnalyzedWorkbook] = []
    for index, preview in enumerate(files):
        if detail_overrides:
            bank_name, account_number = detail_overrides.get(
                index,
                (preview.bank_name or "", preview.account_number or ""),
            )
        else:
            bank_name, account_number = preview.bank_name or "", preview.account_number or ""
        bank_name = bank_name.strip()
        account_number = account_number.strip()
        if not bank_name or not account_number:
            raise ValueError(f"Bank name and account number are required for {preview.filename}.")
        analyzed_workbooks.append(
            load_analyzed_workbook(
                preview.path,
                preview.filename,
                bank_name=bank_name,
                account_number=account_number,
            )
        )

    inflow_headers = merge_headers([workbook.inflows for workbook in analyzed_workbooks], DEFAULT_INFLOW_HEADERS)
    outflow_headers = merge_headers([workbook.outflows for workbook in analyzed_workbooks], DEFAULT_OUTFLOW_HEADERS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(output_path)
    header_format = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1, "align": "center"})
    money_format = workbook.add_format({"num_format": "#,##0.00"})
    date_format = workbook.add_format({"num_format": "dd/mm/yyyy"})

    inflow_count = write_consolidated_sheet(
        workbook,
        "Consolidated Inflows",
        analyzed_workbooks,
        "inflows",
        inflow_headers,
        header_format,
        money_format,
        date_format,
    )
    outflow_count = write_consolidated_sheet(
        workbook,
        "Consolidated Outflows",
        analyzed_workbooks,
        "outflows",
        outflow_headers,
        header_format,
        money_format,
        date_format,
    )
    workbook.close()

    return ConsolidationResult(
        output_path=output_path,
        file_count=len(analyzed_workbooks),
        inflow_count=inflow_count,
        outflow_count=outflow_count,
        files=[workbook.preview for workbook in analyzed_workbooks],
    )


def load_analyzed_workbook(
    path: Path,
    filename: str,
    *,
    bank_name: str,
    account_number: str,
) -> AnalyzedWorkbook:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        _, _, parser_name = extract_metadata(workbook)
        inflows = read_transaction_sheet(workbook, "Inflows")
        outflows = read_transaction_sheet(workbook, "Outflows")
        preview = WorkbookPreview(
            filename=filename,
            path=path,
            bank_name=bank_name,
            account_number=account_number,
            parser_name=parser_name,
            inflow_count=len(inflows.rows),
            outflow_count=len(outflows.rows),
        )
    finally:
        workbook.close()

    return AnalyzedWorkbook(preview=preview, inflows=inflows, outflows=outflows)


def extract_metadata(workbook) -> tuple[str | None, str | None, str | None]:
    bank_name = None
    account_number = None
    parser_name = None

    if "Analysis" in workbook.sheetnames:
        sheet = workbook["Analysis"]
        for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
            if not row:
                continue
            label = normalize_label(row[0])
            value = clean_text_value(row[1] if len(row) > 1 else None)
            if not label or not value:
                continue
            if label in {"bank", "bank name"}:
                bank_name = value
            elif label == "account number":
                account_number = value
            elif label == "parser":
                parser_name = value

    if bank_name and bank_name.strip().lower().startswith("adaptive"):
        bank_name = None
    if not bank_name:
        bank_name = bank_name_from_parser(parser_name)

    return bank_name, account_number, parser_name


def read_transaction_sheet(workbook, sheet_name: str) -> SheetData:
    if sheet_name not in workbook.sheetnames:
        return SheetData(headers=[], rows=[])

    sheet = workbook[sheet_name]
    row_iterator = sheet.iter_rows(values_only=True)
    header_row = next(row_iterator, None)
    if not header_row:
        return SheetData(headers=[], rows=[])

    headers = [clean_header(value) for value in header_row]
    rows: list[dict[str, Any]] = []
    for excel_row_number, row in enumerate(row_iterator, start=2):
        if excel_row_number == 2:
            continue
        if is_blank_row(row):
            continue
        row_values: dict[str, Any] = {}
        for index, header in enumerate(headers):
            normalized = normalize_header(header)
            if not normalized:
                continue
            row_values[normalized] = row[index] if index < len(row) else None
        if any(value not in (None, "") for value in row_values.values()):
            rows.append(row_values)

    return SheetData(headers=[header for header in headers if header], rows=rows)


def write_consolidated_sheet(
    workbook,
    sheet_name: str,
    analyzed_workbooks: list[AnalyzedWorkbook],
    sheet_attribute: str,
    headers: list[tuple[str, str]],
    header_format,
    money_format,
    date_format,
) -> int:
    sheet = workbook.add_worksheet(sheet_name)
    output_headers = with_bank_account_header(headers)
    for column_index, (_, header) in enumerate(output_headers):
        sheet.write(0, column_index, header, header_format)

    output_rows: list[tuple[str, dict[str, Any]]] = []
    for analyzed in analyzed_workbooks:
        source = getattr(analyzed, sheet_attribute)
        bank_account_label = analyzed.preview.bank_account_label
        for row in source.rows:
            output_rows.append((bank_account_label, row))

    for row_index, (bank_account_label, row) in enumerate(output_rows, start=1):
        for column_index, (normalized_header, _) in enumerate(output_headers):
            if normalized_header == BANK_ACCOUNT_KEY:
                sheet.write(row_index, column_index, bank_account_label)
                continue
            value = row.get(normalized_header)
            write_cell(sheet, row_index, column_index, value, money_format, date_format)

    sheet.freeze_panes(1, 0)
    for column_index, (normalized_header, display) in enumerate(output_headers):
        width = 18
        if normalized_header == "DESCRIPTION":
            width = 72
        elif normalized_header in {"DATE", "TRANSACTION DATE", "VALUE DATE"}:
            width = 14
        elif normalized_header == BANK_ACCOUNT_KEY:
            width = 30
        elif normalized_header == "CLASSIFICATION":
            width = 22
        sheet.set_column(column_index, column_index, width)

    return len(output_rows)


def with_bank_account_header(headers: list[tuple[str, str]]) -> list[tuple[str, str]]:
    output_headers = list(headers)
    insertion_index = next(
        (index for index, (normalized, _) in enumerate(output_headers) if normalized == "CLASSIFICATION"),
        None,
    )
    if insertion_index is None:
        insertion_index = next(
            (index + 1 for index, (normalized, _) in enumerate(output_headers) if normalized == "DESCRIPTION"),
            len(output_headers),
        )
    output_headers.insert(insertion_index, (BANK_ACCOUNT_KEY, BANK_ACCOUNT_HEADER))
    return output_headers


def merge_headers(sheet_data_items: list[SheetData], fallback_headers: list[str]) -> list[tuple[str, str]]:
    merged: dict[str, str] = {}
    for sheet_data in sheet_data_items:
        for header in sheet_data.headers:
            normalized = normalize_header(header)
            if normalized and normalized not in merged:
                merged[normalized] = header

    if not merged:
        for header in fallback_headers:
            merged[normalize_header(header)] = header

    return list(merged.items())


def bank_name_from_parser(parser_name: str | None) -> str | None:
    if not parser_name:
        return None
    normalized_parser = parser_name.strip().lower()
    if normalized_parser.startswith("adaptive"):
        return None
    for marker, bank_name in PARSER_BANK_NAMES.items():
        if marker in normalized_parser:
            return bank_name
    return None


def bank_name_from_filename(filename: str) -> str | None:
    normalized = filename.upper()
    hints = (
        ("STANBIC", "Stanbic IBTC"),
        ("KEYSTONE", "Keystone Bank"),
        ("WISDOM KWATI", "Keystone Bank"),
        ("ZENITH", "Zenith Bank"),
        ("UBA", "UBA"),
        ("WEMA", "Wema Bank"),
        ("MONIEPOINT", "Moniepoint"),
        ("OPAY", "OPay"),
    )
    return next((bank for marker, bank in hints if marker in normalized), None)


def write_cell(sheet, row: int, column: int, value: Any, money_format, date_format) -> None:
    if value is None:
        return
    if isinstance(value, datetime):
        sheet.write_datetime(row, column, value, date_format)
        return
    if isinstance(value, date):
        sheet.write_datetime(row, column, datetime.combine(value, datetime.min.time()), date_format)
        return
    if isinstance(value, Decimal):
        sheet.write_number(row, column, float(value), money_format)
        return
    if is_number(value):
        sheet.write_number(row, column, float(value), money_format)
        return
    sheet.write(row, column, value)


def is_totals_or_blank_row(row: tuple[Any, ...]) -> bool:
    if is_blank_row(row):
        return True
    return any(str(value).strip().upper() == "TOTAL" for value in row if value is not None)


def is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(value in (None, "") for value in row)


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def clean_text_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def clean_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def normalize_label(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").strip().upper().split())
