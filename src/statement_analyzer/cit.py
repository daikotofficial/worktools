from __future__ import annotations

from dataclasses import dataclass, replace
from decimal import Decimal, InvalidOperation
import re
from pathlib import Path
from uuid import uuid4

import pdfplumber
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


INPUT_SHEETS = {
    "USER_INPUT",
    "INCOME_STATEMENT",
    "STATEMENT_OF_FINANCIAL_POSITION",
    "PROFIT_ADJUSTMENT",
    "OTHER_INCOME",
    "ADMINISTRATIVE_EXPENSES",
}


@dataclass(frozen=True, slots=True)
class CitTargetCell:
    sheet: str
    cell: str
    label: str
    section: str

    @property
    def key(self) -> str:
        return f"{self.sheet}!{self.cell}"


@dataclass(frozen=True, slots=True)
class CitSourceAmount:
    label: str
    amount: Decimal
    section: str
    page: int
    raw_row: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class CitMappingProposal:
    source: CitSourceAmount
    target: CitTargetCell
    confidence: float
    reason: str


@dataclass(frozen=True, slots=True)
class CitValidationCheck:
    label: str
    expected: Decimal | None
    actual: Decimal | None
    difference: Decimal | None
    status: str


@dataclass(frozen=True, slots=True)
class CitAnalysisResult:
    token: str
    template_filename: str
    afs_filename: str
    targets: list[CitTargetCell]
    sources: list[CitSourceAmount]
    proposals: list[CitMappingProposal]
    unassigned_sources: list[CitSourceAmount]
    checks: list[CitValidationCheck]


def analyze_cit_template(
    template_path: Path,
    afs_path: Path,
    *,
    template_filename: str,
    afs_filename: str,
) -> CitAnalysisResult:
    targets = discover_template_targets(template_path)
    sources = extract_afs_amounts(afs_path)
    accounting_year = extract_accounting_year(
        template_filename=template_filename,
        afs_filename=afs_filename,
        afs_path=afs_path,
    )
    if accounting_year is not None:
        sources.insert(
            0,
            CitSourceAmount(
                label="Accounting year",
                amount=Decimal(accounting_year),
                section="user_input",
                page=0,
                raw_row=("Accounting year", str(accounting_year)),
            ),
        )
    proposals = propose_mappings(sources, targets)
    unassigned_sources = build_unassigned_sources(sources, proposals)
    checks = build_validation_checks(sources, proposals)
    return CitAnalysisResult(
        token=uuid4().hex,
        template_filename=template_filename,
        afs_filename=afs_filename,
        targets=targets,
        sources=sources,
        proposals=proposals,
        unassigned_sources=unassigned_sources,
        checks=checks,
    )


def discover_template_targets(template_path: Path) -> list[CitTargetCell]:
    workbook = load_workbook(template_path, data_only=False)
    targets: list[CitTargetCell] = []
    for worksheet in workbook.worksheets:
        if worksheet.title not in INPUT_SHEETS:
            continue
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.protection.locked is not False:
                    continue
                targets.append(
                    CitTargetCell(
                        sheet=worksheet.title,
                        cell=cell.coordinate,
                        label=target_label_for_cell(worksheet, cell.row, cell.column),
                        section=target_section_for_cell(worksheet.title, cell.row),
                    )
                )
    return targets


def target_label_for_cell(worksheet: Worksheet, row: int, column: int) -> str:
    candidates = []
    for offset in range(1, 4):
        value = worksheet.cell(row, max(1, column - offset)).value
        if isinstance(value, str) and value.strip():
            candidates.append(clean_text(value))
    if candidates:
        return candidates[0]
    value = worksheet.cell(row, column).value
    return clean_text(str(value)) if value is not None else f"{worksheet.title}!{worksheet.cell(row, column).coordinate}"


def target_section_for_cell(sheet: str, row: int) -> str:
    if sheet == "INCOME_STATEMENT":
        if 7 <= row <= 27:
            return "revenue"
        if 66 <= row <= 97:
            return "cost_of_sales"
        if row in {113, 114, 116, 119, 120, 121, 122, 123, 124, 125, 127}:
            return "profit_or_loss"
    if sheet == "STATEMENT_OF_FINANCIAL_POSITION":
        if row <= 43:
            return "non_current_assets"
        if row <= 69:
            return "current_assets"
        if row <= 110:
            return "liabilities"
        return "equity"
    if sheet == "OTHER_INCOME":
        return "other_income"
    if sheet == "ADMINISTRATIVE_EXPENSES":
        return "administrative_expenses"
    if sheet == "PROFIT_ADJUSTMENT":
        return "profit_adjustment"
    if sheet == "USER_INPUT":
        return "user_input"
    return sheet.lower()


def extract_afs_amounts(pdf_path: Path) -> list[CitSourceAmount]:
    sources: list[CitSourceAmount] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            section = "unknown"
            text = page.extract_text() or ""
            for line in text.splitlines():
                cleaned_line = clean_text(line)
                if not cleaned_line:
                    continue
                detected = detect_section(normalized_text(cleaned_line))
                if detected:
                    section = detected
                source = source_amount_from_text_line(cleaned_line, section=section, page=page_number)
                if source is not None:
                    sources.append(source)
            tables = page.extract_tables() or []
            for table in tables:
                for row in table:
                    row_values = tuple(clean_text(value or "") for value in row)
                    source = source_amount_from_row(row_values, section=section, page=page_number)
                    if source is not None:
                        sources.append(source)
                        continue
                    joined = normalized_text(" ".join(row_values))
                    detected = detect_section(joined)
                    if detected:
                        section = detected
    return compact_sources(sources)


def detect_section(row_text: str) -> str | None:
    if not row_text:
        return None
    if row_text.startswith("STATEMENT OF CASH FLOW") or row_text.startswith("STATEMENTS OF CASH FLOW"):
        return None
    section_markers = (
        ("STATEMENT OF FINANCIAL POSITION", "financial_position"),
        ("PROFIT AND LOSS", "profit_or_loss"),
        ("INCOME STATEMENT", "profit_or_loss"),
        ("STATEMENT OF CHANGES IN EQUITY", "equity_statement"),
        ("REVENUE", "revenue_note"),
        ("DIRECT COST", "cost_note"),
        ("COST OF SALES", "cost_note"),
        ("OTHER INCOME", "other_income_note"),
        ("OTHER OPERATING INCOME", "other_income_note"),
        ("CASH & CASH EQUIVALENT", "cash_note"),
        ("CASH AND CASH EQUIVALENT", "cash_note"),
        ("TRADE AND OTHER REC", "receivables_note"),
        ("TRADE AND OTHER PAYABLE", "payables_note"),
        ("CURRENT LIABILITY", "payables_note"),
        ("CURRENT LIABILITIES", "payables_note"),
        ("CREDITORS & ACCRUAL", "payables_note"),
        ("PAYABLES", "payables_note"),
        ("ADMINISTRATIVE EXPENSE", "administrative_expenses_note"),
        ("SELLING & ADMINISTRATIVE EXPENSE", "administrative_expenses_note"),
        ("GENERAL AND ADMINISTRATIVE", "administrative_expenses_note"),
        ("TAX COMPUTATION", "tax_computation"),
        ("TAXATION PROVISION", "taxation_note"),
        ("INCOME TAX", "taxation_note"),
        ("CURRENT TAX", "taxation_note"),
        ("PROPERTY PLANT AND EQUIPMENT", "ppe_note"),
        ("PROPERTY, PLANT", "ppe_note"),
        ("FUNDED BY", "equity_note"),
        ("SHARE CAPITAL", "equity_note"),
        ("STATEMENT OF CHANGES IN EQUITY", "equity_statement"),
    )
    for marker, section in section_markers:
        if marker in row_text:
            return section
    return None


def source_amount_from_row(row: tuple[str, ...], *, section: str, page: int) -> CitSourceAmount | None:
    if not row:
        return None
    label_parts: list[str] = []
    amount: Decimal | None = None
    for value in row:
        if not value:
            continue
        candidate = parse_money(value)
        if candidate is not None and not is_small_ungrouped_number(value, candidate):
            amount = candidate
            break
        if not looks_like_note_number(value):
            label_parts.append(value)
    if amount is None:
        return None
    label = clean_label(" ".join(label_parts))
    if not label or label.upper() in {"2025", "NOTES", "NOTE"}:
        return None
    return CitSourceAmount(label=label, amount=amount, section=section, page=page, raw_row=row)


def source_amount_from_text_line(line: str, *, section: str, page: int) -> CitSourceAmount | None:
    if not line or should_skip_text_line(line):
        return None
    prepared = prepare_money_text(line)
    matches = list(re.finditer(r"\(?-?\d[\d,]*(?:\.\d+)?\)?", prepared))
    if not matches:
        return None

    candidates: list[tuple[re.Match[str], Decimal]] = []
    for match in matches:
        raw = match.group(0)
        amount = parse_money(raw)
        if amount is None:
            continue
        if is_probable_note_or_year(raw, amount, match, matches):
            continue
        candidates.append((match, amount))
    if not candidates:
        return None

    match, amount = candidates[0]
    label = clean_label(prepared[: match.start()])
    label = re.sub(r"\b[NS]CE\b$", "", label, flags=re.IGNORECASE)
    label = re.sub(r"\bSOCE\b$", "", label, flags=re.IGNORECASE)
    label = re.sub(r"\bFTY\b$", "", label, flags=re.IGNORECASE)
    label = re.sub(r"\bDTY\b$", "", label, flags=re.IGNORECASE)
    label = re.sub(r"\bPPE\b$", "", label, flags=re.IGNORECASE)
    label = clean_text(label)
    if not label or normalized_text(label) in {"N", "NOTES", "TOTAL N", "2025", "2024"}:
        return None
    return CitSourceAmount(label=label, amount=amount, section=section, page=page, raw_row=(line,))


def should_skip_text_line(line: str) -> bool:
    upper = normalized_text(line)
    if upper.startswith("PAGE "):
        return True
    if "AUDITED FINANCIAL STATEMENT" in upper or "FINANCIAL STATEMENTS" in upper:
        return True
    if "FOR THE YEAR ENDED" in upper or "YEAR ENDED" in upper:
        return True
    if re.search(r"\b20\d{2}\s+20\d{2}\b", upper) and not re.search(r"\d[\d,]{3}", upper):
        return True
    if "%" in upper and not re.search(r"\d[\d,]{3}", upper):
        return True
    return False


def prepare_money_text(line: str) -> str:
    prepared = clean_text(line)
    prepared = prepared.replace("₦", "N")
    for _ in range(3):
        prepared = re.sub(r"(?<=\d)\s+,(?=\d)", ",", prepared)
        prepared = re.sub(r"(?<=\d),\s+(?=\d)", ",", prepared)
        prepared = re.sub(r"(?<![\d,])(\d)\s+(\d)\s+(?=\d{2},\d{3}(?:,\d{3})*\b)", r"\1 \2", prepared)
        prepared = re.sub(r"(?<![\d,])(\d)\s+(?=0,?\d{3}(?:,\d{3})*\b|00,\d{3}(?:,\d{3})*\b)", r"\1", prepared)
    prepared = re.sub(r"\(\s+", "(", prepared)
    prepared = re.sub(r"\s+\)", ")", prepared)
    return prepared


def is_probable_note_or_year(
    raw: str,
    amount: Decimal,
    match: re.Match[str],
    matches: list[re.Match[str]],
) -> bool:
    bare = raw.replace(",", "").strip("()")
    has_grouping = "," in raw or "." in raw or "(" in raw
    if bare in {"2024", "2025", "2026"}:
        return True
    if is_small_ungrouped_number(raw, amount):
        return True
    if match.end() < 5 and not has_grouping and amount.copy_abs() <= 99:
        return True
    return False


def is_small_ungrouped_number(raw: str, amount: Decimal) -> bool:
    return amount.copy_abs() <= 99 and not any(marker in raw for marker in (",", ".", "(", ")"))


def compact_sources(sources: list[CitSourceAmount]) -> list[CitSourceAmount]:
    seen: set[tuple[str, str, Decimal, int]] = set()
    compacted: list[CitSourceAmount] = []
    for source in sources:
        key = (normalized_text(source.label), source.section, source.amount, source.page)
        if key in seen:
            continue
        seen.add(key)
        compacted.append(source)
    return compacted


def propose_mappings(sources: list[CitSourceAmount], targets: list[CitTargetCell]) -> list[CitMappingProposal]:
    targets_by_key = {target.key: target for target in targets}
    proposals: list[CitMappingProposal] = []
    used_targets: set[str] = set()

    for source in sources:
        target_key, confidence, reason = choose_target(source)
        if not target_key or target_key not in targets_by_key or target_key in used_targets:
            continue
        target = targets_by_key[target_key]
        source_for_proposal = normalize_source_for_target(source, target)
        used_targets.add(target_key)
        proposals.append(
            CitMappingProposal(
                source=source_for_proposal,
                target=target,
                confidence=confidence,
                reason=reason,
            )
        )
    return proposals


def normalize_source_for_target(source: CitSourceAmount, target: CitTargetCell) -> CitSourceAmount:
    should_be_positive = (
        target.sheet == "ADMINISTRATIVE_EXPENSES"
        or target.section in {"cost_of_sales", "profit_adjustment"}
        or target.section in {"non_current_assets", "current_assets", "liabilities"}
        or target.key in {
            "INCOME_STATEMENT!D116",
            "STATEMENT_OF_FINANCIAL_POSITION!D73",
            "STATEMENT_OF_FINANCIAL_POSITION!D74",
            "STATEMENT_OF_FINANCIAL_POSITION!D76",
            "STATEMENT_OF_FINANCIAL_POSITION!D78",
            "STATEMENT_OF_FINANCIAL_POSITION!D89",
            "STATEMENT_OF_FINANCIAL_POSITION!D91",
            "STATEMENT_OF_FINANCIAL_POSITION!D95",
            "STATEMENT_OF_FINANCIAL_POSITION!D96",
            "STATEMENT_OF_FINANCIAL_POSITION!D110",
        }
    )
    if should_be_positive and source.amount < 0:
        return replace(source, amount=abs(source.amount))
    return source


def choose_target(source: CitSourceAmount) -> tuple[str | None, float, str]:
    text = normalized_text(source.label)
    section = source.section

    if should_ignore_source_label(text):
        return None, 0.0, ""

    if section == "user_input" and "ACCOUNTING YEAR" in text:
        return "USER_INPUT!D7", 0.96, "accounting year detected from template/AFS naming"

    if is_total_label(text) and section == "revenue_note":
        return "INCOME_STATEMENT!D27", 0.86, "total revenue in AFS note mapped to Others revenue pending sector classification"
    if section == "revenue_note":
        if any(marker in text for marker in ("AGO", "PMS", "OIL", "GAS")):
            return "INCOME_STATEMENT!D23", 0.74, "oil/gas related revenue mapped to oil and gas"
        if "CONTRACT" in text:
            return "INCOME_STATEMENT!D27", 0.62, "contract revenue proposed as other revenue for review"
    if is_total_label(text) and section == "cost_note":
        return "INCOME_STATEMENT!D84", 0.86, "total cost/direct cost in AFS note mapped to Other Direct costs"
    if is_total_label(text) and section == "other_income_note":
        return "OTHER_INCOME!D17", 0.82, "total other income mapped to Others"

    if section == "profit_or_loss":
        if "INCOME TAX" in text or "PROVISION FOR TAX" in text:
            return "INCOME_STATEMENT!D116", 0.72, "income tax charge mapped to taxation"
        if "PROFIT AFTER TAX" in text or "PROFIT FROM CONTINUING" in text:
            return "INCOME_STATEMENT!D127", 0.58, "profit after tax proposed as retained earnings movement for review"
        return None, 0.0, ""

    if section in {"financial_position", "receivables_note", "cash_note", "ppe_note"}:
        if "PROPERTY" in text or "PPE" in text or "NON-CURRENT ASSETS" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D8", 0.82, "PPE/non-current asset amount mapped to property, plant and equipment"
        if "INVESTMENT" in text and "STAFF" not in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D19", 0.68, "investment asset proposed as other non-current assets"
        if "INVENTOR" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D46", 0.82, "inventory mapped to inventories"
        if "WHT" in text or "WITHHOLDING" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D51", 0.88, "withholding tax credit mapped to WHT receivables"
        if "TRADE REC" in text or "OTHER RECEIV" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D48", 0.9, "trade receivables mapped to trade receivables"
        if "DIRECTOR" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D62", 0.78, "director account mapped to other current assets - Others"
        if "CASH" in text or "BANK BALANCE" in text or "CASH AT BANK" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D65", 0.9, "cash and cash equivalents mapped to cash at bank"
        if "PREPAY" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D58", 0.76, "prepayments mapped to prepaid expenses"

    if section in {"financial_position", "equity_note"}:
        if "SHARE CAPITAL" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D112", 0.93, "share capital mapped to issued share capital"
        if "RETAINED EARN" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D114", 0.93, "retained earnings mapped to retained earnings"
        if "PARTNER" in text or "RESERVE" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D124", 0.7, "other equity funding/reserve proposed as other reserves"
        if "DIRECTOR" in text and "FUND" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D125", 0.72, "director/shareholder fund proposed as equity contribution"

    if section in {"financial_position", "payables_note", "equity_note", "taxation_note"}:
        if "LONG TERM" in text and ("LOAN" in text or "LIABIL" in text or "BORROW" in text):
            return "STATEMENT_OF_FINANCIAL_POSITION!D95", 0.74, "long-term liability mapped to long-term debt"
        if "BANK LOAN" in text or "BORROW" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D96", 0.74, "bank loan/borrowings mapped to long-term borrowings"
        if "DIRECTOR" in text and ("LOAN" in text or "CURRENT ACCOUNT" in text):
            return "STATEMENT_OF_FINANCIAL_POSITION!D76", 0.7, "director payable mapped to other payables"
        if "TAX PROVISION" in text or "PROVISION FOR TAX" in text or "CURRENT TAX" in text or "INCOME TAX" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D89", 0.72, "tax provision mapped to income taxes payable"
        if "ACCRUAL" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D78", 0.78, "accruals mapped to accruals"
        if "TRADE PAYABLE" in text or "ACCOUNT PAYABLE" in text or "PAYABLES" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D73", 0.76, "payables mapped to trade payables"

    if section == "equity_note":
        if "SHARE CAPITAL" in text:
            return "STATEMENT_OF_FINANCIAL_POSITION!D112", 0.9, "funding note share capital mapped to issued share capital"

    if section == "administrative_expenses_note":
        return admin_expense_target(text)
    if section == "cost_note":
        direct_target = direct_cost_target(text)
        if direct_target[0]:
            return direct_target
        admin_target = admin_expense_target(text)
        if admin_target[0]:
            return admin_target

    return None, 0.0, ""


def should_ignore_source_label(text: str) -> bool:
    if not text:
        return True
    ignored_exact = {
        "TOTAL ASSETS",
        "TOTAL ASSET",
        "TOTAL CURRENT ASSETS",
        "TOTAL CURRENT ASSET",
        "TOTAL LIABILITIES",
        "TOTAL LIABILITY",
        "TOTAL EQUITY",
        "TOTAL CAPITAL & LIABILITIES",
        "TOTAL EQUITY AND LIABILITIES",
        "TOTAL SHAREHOLDER'S EQUITIES",
        "TOTAL SHAREHOLDERS EQUITY",
        "TOTAL COMPREHENSIVE INCOME",
        "GROSS PROFIT",
        "OPERATING PROFIT",
        "PROFIT BEFORE FINANCING AND INCOME TAXES",
        "PROFIT BEFORE INCOME TAXES",
        "PROFIT OR LOSS BEFORE TAX",
    }
    if text in ignored_exact:
        return True
    if text.startswith("STATEMENT OF ") or text.startswith("STATEMENTS OF "):
        return True
    cash_flow_terms = (
        "BEGINNING OF THE YEAR",
        "NET DECREASE",
        "NET (DECREASE",
        "NET INCREASE",
        "NET (INCREASE",
        "INCREASE IN CASH",
        "DECREASE IN CASH",
        "NET CASH",
        "CASH FLOWS",
        "OPERATING ACTIVITIES",
        "INVESTING ACTIVITIES",
        "FINANCING ACTIVITIES",
    )
    if any(term in text for term in cash_flow_terms):
        return True
    if text.startswith("TOTAL ") and any(term in text for term in ("ASSET", "LIABIL", "EQUITY", "CAPITAL")):
        return True
    if "SECTION " in text or "CAMA" in text or "FRC/" in text:
        return True
    return False


def direct_cost_target(text: str) -> tuple[str | None, float, str]:
    if is_total_label(text) or text in {"DIRECT COST", "DIRECT COST:"} or "COST OF SALES" in text or "LESS DIRECT COST" in text:
        return "INCOME_STATEMENT!D84", 0.8, "direct cost/cost of sales mapped to other direct costs"
    if "PURCHASE" in text:
        return "INCOME_STATEMENT!D71", 0.78, "purchases mapped to local purchases"
    if "LOGISTIC" in text:
        return "INCOME_STATEMENT!D84", 0.68, "logistics cost proposed as other direct costs"
    if "DEPRECIATION" in text and "OPERATING" not in text:
        return "INCOME_STATEMENT!D79", 0.68, "direct-cost depreciation mapped to cost of sales depreciation"
    if "CONSTRUCTION" in text or "BUILDING COST" in text:
        return "INCOME_STATEMENT!D81", 0.76, "construction/building cost mapped to direct construction cost"
    if "CONTRACT COST" in text or "SUB-CONTRACT" in text or "SUBCONTRACT" in text:
        return "INCOME_STATEMENT!D84", 0.7, "contract/subcontract cost proposed as other direct costs"
    return None, 0.0, ""


def admin_expense_target(text: str) -> tuple[str | None, float, str]:
    if "GENERAL AND ADMINISTRATIVE" in text or is_total_label(text):
        return None, 0.0, ""
    if "PROFIT" in text or "CONTINUING OPERATIONS" in text:
        return None, 0.0, ""
    if "TAX" in text and not ("AUDIT" in text or "CONSULT" in text):
        return None, 0.0, ""

    rules = (
        (("DEPRECIATION",), "ADMINISTRATIVE_EXPENSES!C13", "depreciation"),
        (("BUSINESS COMPLIANCE", "REGULATORY", "STATUTORY"), "ADMINISTRATIVE_EXPENSES!C15", "government/regulatory cost"),
        (("ADVERTISEMENT", "ADVERTIS"), "ADMINISTRATIVE_EXPENSES!C16", "advertisement and promotion"),
        (("BANK CHARGE",), "ADMINISTRATIVE_EXPENSES!C18", "bank charges"),
        (("INSURANCE",), "ADMINISTRATIVE_EXPENSES!C67", "insurance/statutory cost"),
        (("AUDIT",), "ADMINISTRATIVE_EXPENSES!C21", "audit fees"),
        (("DONATION", "CORPORATE GIFT"), "ADMINISTRATIVE_EXPENSES!C22", "donations/corporate gift"),
        (("ENTERTAIN", "OFFICE CONSUMABLE"), "ADMINISTRATIVE_EXPENSES!C23", "entertainment or consumables requiring review"),
        (("AGO", "DIESEL"), "ADMINISTRATIVE_EXPENSES!C68", "AGO/diesel"),
        (("FUEL", "LUBRICANT", "PMS"), "ADMINISTRATIVE_EXPENSES!C25", "fuel expenses"),
        (("RENT",), "ADMINISTRATIVE_EXPENSES!C34", "rent and rates"),
        (("REPAIR", "MAINTENANCE"), "ADMINISTRATIVE_EXPENSES!C35", "repairs and maintenance"),
        (("TELEPHONE", "COMMUNICATION"), "ADMINISTRATIVE_EXPENSES!C43", "telephone/communication"),
        (("TRANSPORT",), "ADMINISTRATIVE_EXPENSES!C46", "transport of supplies"),
        (("UTILITY", "UTILITIES"), "ADMINISTRATIVE_EXPENSES!C47", "utilities"),
        (("PROFESSIONAL", "CONSULT"), "ADMINISTRATIVE_EXPENSES!C56", "professional fees"),
        (("LODGING", "ROOM", "HOTEL", "ACCOMODATION", "ACCOMMODATION"), "ADMINISTRATIVE_EXPENSES!C20", "rooms/accommodation cost"),
        (("TRAVEL", "TRIP"), "ADMINISTRATIVE_EXPENSES!C69", "travel expenses"),
        (("OFFICE SUPPLIES", "SUPPLIES", "STATIONER", "SATIONAR"), "ADMINISTRATIVE_EXPENSES!C55", "stationeries"),
        (("SALARY", "STAFF COST", "WAGE", "PERSONNEL", "PAYROLL", "EMPLOYEE COST", "EMPLOYEE EXPENSE", "MANPOWER"), "ADMINISTRATIVE_EXPENSES!C4", "salary and wages/personnel cost"),
        (("EMPLOYEE BENEFIT", "STAFF WELFARE", "STAFF BENEFIT"), "ADMINISTRATIVE_EXPENSES!C5", "employee benefit expenses"),
        (("DIRECTOR EMOLUMENT", "DIRECTORS EMOLUMENT"), "ADMINISTRATIVE_EXPENSES!C6", "directors emoluments"),
        (("DIRECTOR EXPENSE", "DIRECTORS EXPENSE"), "ADMINISTRATIVE_EXPENSES!C7", "directors expenses"),
        (("ALLOWANCE",), "ADMINISTRATIVE_EXPENSES!C8", "staff allowances"),
        (("PENSION", "GRATUITY"), "ADMINISTRATIVE_EXPENSES!C11", "pension and gratuity"),
        (("BAD DEBT",), "ADMINISTRATIVE_EXPENSES!C17", "bad debt written off"),
        (("INTEREST",), "ADMINISTRATIVE_EXPENSES!C19", "interest expense"),
        (("POSTAGE", "COURIER"), "ADMINISTRATIVE_EXPENSES!C24", "postages"),
        (("GENERATOR",), "ADMINISTRATIVE_EXPENSES!C26", "generator expenses"),
        (("HEAD OFFICE",), "ADMINISTRATIVE_EXPENSES!C27", "head office cost"),
        (("INTERNET",), "ADMINISTRATIVE_EXPENSES!C28", "internet cost"),
        (("LOSS ON SALE", "DISPOSAL"), "ADMINISTRATIVE_EXPENSES!C29", "loss on sale of PPE"),
        (("MEDICAL", "HEALTHCARE"), "ADMINISTRATIVE_EXPENSES!C30", "medical expenses"),
        (("NEWSPAPER", "PERIODICAL"), "ADMINISTRATIVE_EXPENSES!C31", "newspaper and periodicals"),
        (("OUTSOURC",), "ADMINISTRATIVE_EXPENSES!C32", "outsourcing services"),
        (("RATE", "LEVY"), "ADMINISTRATIVE_EXPENSES!C33", "rates"),
        (("SECURITY",), "ADMINISTRATIVE_EXPENSES!C37", "security expenses"),
        (("SELLING", "DISTRIBUTION"), "ADMINISTRATIVE_EXPENSES!C38", "selling and distribution"),
        (("SERVICE CHARGE",), "ADMINISTRATIVE_EXPENSES!C39", "service charge"),
        (("PUBLIC RELATION",), "ADMINISTRATIVE_EXPENSES!C40", "public relation"),
        (("PUBLICATION",), "ADMINISTRATIVE_EXPENSES!C41", "publication cost"),
        (("PERMIT", "LICENCE", "LICENSE"), "ADMINISTRATIVE_EXPENSES!C42", "permits/licences"),
        (("TRAINING", "DEVELOPMENT"), "ADMINISTRATIVE_EXPENSES!C45", "training and development"),
        (("WEBSITE",), "ADMINISTRATIVE_EXPENSES!C48", "website design and maintenance"),
        (("SOFTWARE",), "ADMINISTRATIVE_EXPENSES!C49", "software maintenance"),
        (("CLEANING", "JANITOR"), "ADMINISTRATIVE_EXPENSES!C50", "cleaning and janitorial services"),
        (("PROMOTION",), "ADMINISTRATIVE_EXPENSES!C51", "corporate promotion"),
        (("FREIGHT",), "ADMINISTRATIVE_EXPENSES!C52", "freight and transport"),
        (("SUBSCRIPTION", "DUES"), "ADMINISTRATIVE_EXPENSES!C53", "subscription and dues"),
        (("SPONSOR",), "ADMINISTRATIVE_EXPENSES!C54", "sponsorship"),
        (("FINE",), "ADMINISTRATIVE_EXPENSES!C57", "fines"),
        (("PENALT",), "ADMINISTRATIVE_EXPENSES!C58", "penalties"),
        (("AMCON",), "ADMINISTRATIVE_EXPENSES!C59", "AMCON charges"),
        (("FOREIGN EXCHANGE LOSS", "FX LOSS"), "ADMINISTRATIVE_EXPENSES!C60", "realized foreign exchange loss"),
        (("CSR", "SOCIAL RESPONSIB"), "ADMINISTRATIVE_EXPENSES!C61", "corporate social responsibility"),
        (("ROYALT",), "ADMINISTRATIVE_EXPENSES!C62", "royalties"),
        (("ELECTRICITY", "POWER", "WATER"), "ADMINISTRATIVE_EXPENSES!C63", "electricity and power"),
        (("HEALTH", "SAFETY", "ENVIRONMENT"), "ADMINISTRATIVE_EXPENSES!C65", "health, safety and environmental expenses"),
        (("STATUTORY CHARGE", "STATUTORY LEVY"), "ADMINISTRATIVE_EXPENSES!C67", "other statutory charges and levies"),
    )
    for markers, target, reason in rules:
        if any(marker in text for marker in markers):
            return target, 0.78, reason
    if not is_total_label(text):
        return "ADMINISTRATIVE_EXPENSES!C70", 0.45, "unmatched administrative expense placed in Others for review"
    return None, 0.0, ""


def build_unassigned_sources(
    sources: list[CitSourceAmount],
    proposals: list[CitMappingProposal],
) -> list[CitSourceAmount]:
    assigned = {
        (
            normalized_text(proposal.source.label),
            proposal.source.section,
            proposal.source.amount,
            proposal.source.page,
        )
        for proposal in proposals
    }
    review_sections = {
        "revenue_note",
        "cost_note",
        "other_income_note",
        "administrative_expenses_note",
        "taxation_note",
    }
    unassigned: list[CitSourceAmount] = []
    for source in sources:
        key = (normalized_text(source.label), source.section, source.amount, source.page)
        if key in assigned:
            continue
        if source.section not in review_sections:
            continue
        label = normalized_text(source.label)
        if is_total_label(label):
            continue
        if "GENERAL AND ADMINISTRATIVE" in label:
            continue
        if "PROFIT" in label:
            continue
        unassigned.append(source)
    return unassigned


def build_validation_checks(
    sources: list[CitSourceAmount],
    proposals: list[CitMappingProposal],
) -> list[CitValidationCheck]:
    proposed_by_target = {proposal.target.key: proposal.source.amount for proposal in proposals}
    checks: list[CitValidationCheck] = []
    add_check(checks, "Revenue", find_source_total(sources, "revenue_note"), proposed_by_target.get("INCOME_STATEMENT!D27"))
    add_check(checks, "Cost/direct cost", find_source_total(sources, "cost_note"), proposed_by_target.get("INCOME_STATEMENT!D84"))
    add_check(checks, "Other income", find_source_total(sources, "other_income_note"), proposed_by_target.get("OTHER_INCOME!D17"))
    add_check(checks, "Administrative expenses", find_source_total(sources, "administrative_expenses_note"), sum_targets(proposed_by_target, "ADMINISTRATIVE_EXPENSES!C"))
    add_check(
        checks,
        "Financial position balance",
        sum_targets(proposed_by_target, "STATEMENT_OF_FINANCIAL_POSITION!D", cells={"D8", "D10", "D11"}),
        None,
        status="review",
    )
    return checks


def add_check(
    checks: list[CitValidationCheck],
    label: str,
    expected: Decimal | None,
    actual: Decimal | None,
    *,
    status: str | None = None,
) -> None:
    difference = (actual - expected) if expected is not None and actual is not None else None
    if status is None:
        if expected is None or actual is None:
            status = "review"
        elif abs(difference or Decimal("0")) <= Decimal("1"):
            status = "matched"
        else:
            status = "mismatch"
    checks.append(CitValidationCheck(label=label, expected=expected, actual=actual, difference=difference, status=status))


def find_source_total(sources: list[CitSourceAmount], section: str) -> Decimal | None:
    totals = [source.amount for source in sources if source.section == section and is_total_label(normalized_text(source.label))]
    return totals[-1] if totals else None


def extract_accounting_year(*, template_filename: str, afs_filename: str, afs_path: Path) -> int | None:
    for value in (template_filename, afs_filename):
        match = re.search(r"(?<!\d)(20\d{2})(?!\d)", value)
        if match:
            return int(match.group(1))

    try:
        with pdfplumber.open(afs_path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages[:3])
    except Exception:
        return None

    patterns = (
        r"YEAR ENDED\s+31\s+DECEMBER\s+(20\d{2})",
        r"YEAR ENDED\s+31ST\s+DEC\s+(20\d{2})",
        r"FOR THE YEAR ENDED.*?(20\d{2})",
    )
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return int(match.group(1))
    return None


def sum_targets(proposed_by_target: dict[str, Decimal], prefix: str, *, cells: set[str] | None = None) -> Decimal | None:
    values = []
    for target, amount in proposed_by_target.items():
        if not target.startswith(prefix):
            continue
        if cells is not None and target.split("!", 1)[1] not in cells:
            continue
        values.append(amount)
    return sum(values, Decimal("0")) if values else None


def populate_cit_workbook(
    template_path: Path,
    output_path: Path,
    mappings: dict[str, Decimal],
) -> None:
    targets = {target.key for target in discover_template_targets(template_path)}
    workbook = load_workbook(template_path, data_only=False)
    for key, amount in mappings.items():
        if key not in targets:
            raise ValueError(f"{key} is not an editable NRS template cell.")
        sheet_name, cell = key.split("!", 1)
        workbook[sheet_name][cell] = float(amount) if amount == amount.to_integral_value() else float(amount)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_review_mappings(raw_items: list[tuple[str, str]]) -> dict[str, Decimal]:
    mappings: dict[str, Decimal] = {}
    for target, amount_text in raw_items:
        target = target.strip()
        if not target:
            continue
        amount = parse_money(amount_text)
        if amount is None:
            continue
        mappings[target] = amount
    return mappings


def parse_money(value: str | None) -> Decimal | None:
    if value is None:
        return None
    cleaned = clean_text(value).upper()
    if not cleaned or cleaned in {"-", "--", "NIL", "NONE", "₦"}:
        return None
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.replace("₦", "").replace(",", "").replace(" ", "")
    cleaned = cleaned.strip("()")
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    if not cleaned or cleaned in {"-", ".", "-."}:
        return None
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -amount if negative and amount > 0 else amount


def clean_label(value: str) -> str:
    cleaned = clean_text(value)
    cleaned = re.sub(r"^\d+\s+", "", cleaned)
    cleaned = re.sub(r"\s+\d+$", "", cleaned)
    cleaned = re.sub(r"\bNotes?\b", "", cleaned, flags=re.IGNORECASE)
    return clean_text(cleaned)


def clean_text(value: str | None) -> str:
    return " ".join((value or "").replace("\n", " ").replace("\x00", " ").split())


def normalized_text(value: str | None) -> str:
    return clean_text(value).upper()


def looks_like_note_number(value: str) -> bool:
    return bool(re.fullmatch(r"\d+|SOCE", value.strip(), flags=re.IGNORECASE))


def is_total_label(text: str) -> bool:
    return text == "TOTAL" or text.startswith("TOTAL ")
