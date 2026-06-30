from __future__ import annotations

import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from statement_analyzer.consolidation import (
    BANK_ACCOUNT_HEADER,
    consolidate_analyzed_workbooks,
    inspect_analyzed_workbook,
)
from statement_analyzer.exporters.excel import ExcelExporter
from statement_analyzer.models import ClassifiedTransaction, StatementAnalysis, StatementMetadata, Transaction


class ConsolidationTests(unittest.TestCase):
    def test_consolidates_inflows_and_outflows_from_analyzed_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base_path = Path(temp_dir)
            first_path = base_path / "zenith_ANALYZED.xlsx"
            second_path = base_path / "wema_ANALYZED.xlsx"
            output_path = base_path / "consolidated.xlsx"

            write_analyzed_workbook(
                first_path,
                parser_name="zenith-style",
                account_number="1010191819",
                inflow_description="Payment from ALPHA LTD",
                outflow_description="Transfer to supplier",
            )
            write_analyzed_workbook(
                second_path,
                parser_name="wema",
                account_number="2020202020",
                inflow_description="Payment from BETA LTD",
                outflow_description="Office supplies",
            )

            previews = [
                inspect_analyzed_workbook(first_path),
                inspect_analyzed_workbook(second_path),
            ]
            result = consolidate_analyzed_workbooks(previews, output_path)

            self.assertEqual(result.file_count, 2)
            self.assertEqual(result.inflow_count, 2)
            self.assertEqual(result.outflow_count, 2)
            self.assertEqual(previews[0].bank_name, "Zenith Bank")
            self.assertFalse(previews[0].needs_manual_details)

            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Consolidated Inflows", "Consolidated Outflows"])
            inflows = workbook["Consolidated Inflows"]
            outflows = workbook["Consolidated Outflows"]

            inflow_headers = [cell.value for cell in inflows[1]]
            outflow_headers = [cell.value for cell in outflows[1]]
            self.assertEqual(inflow_headers[:6], ["DATE", "DESCRIPTION", "DEBIT", "CREDIT", BANK_ACCOUNT_HEADER, "CLASSIFICATION"])
            self.assertEqual(outflow_headers[:7], ["DATE", "DESCRIPTION", "DEBIT", "CONFIRM", "DIFF", BANK_ACCOUNT_HEADER, "CLASSIFICATION"])
            self.assertNotEqual(inflows["A2"].value, "TOTAL")
            self.assertNotEqual(outflows["A2"].value, "TOTAL")
            self.assertEqual(inflows["E2"].value, "Zenith Bank 1010191819")
            self.assertEqual(inflows["E3"].value, "Wema Bank 2020202020")
            self.assertEqual(outflows["F2"].value, "Zenith Bank 1010191819")
            self.assertEqual(outflows["F3"].value, "Wema Bank 2020202020")

            inflow_classification_index = inflow_headers.index("CLASSIFICATION")
            outflow_classification_index = outflow_headers.index("CLASSIFICATION")
            self.assertGreater(inflow_headers.index("SALES"), inflow_classification_index)
            self.assertGreater(outflow_headers.index("GOODS"), outflow_classification_index)

            inflow_sales_column = inflow_headers.index("SALES") + 1
            outflow_goods_column = outflow_headers.index("GOODS") + 1
            self.assertEqual(inflows.cell(row=2, column=inflow_sales_column).value, 250000)
            self.assertEqual(outflows.cell(row=2, column=outflow_goods_column).value, 100000)

    def test_manual_bank_details_can_fill_detection_gaps(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            base_path = Path(temp_dir)
            input_path = base_path / "adaptive_ANALYZED.xlsx"
            output_path = base_path / "consolidated.xlsx"

            write_analyzed_workbook(
                input_path,
                parser_name="adaptive-unknown",
                account_number="3030303030",
                inflow_description="Client lodgement",
                outflow_description="Vendor payout",
            )

            preview = inspect_analyzed_workbook(input_path)
            self.assertIsNone(preview.bank_name)
            self.assertTrue(preview.needs_manual_details)

            result = consolidate_analyzed_workbooks(
                [preview],
                output_path,
                detail_overrides={0: ("Manual Bank", "3030303030")},
            )

            self.assertEqual(result.files[0].bank_account_label, "Manual Bank 3030303030")
            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook["Consolidated Inflows"]["E2"].value, "Manual Bank 3030303030")


def write_analyzed_workbook(
    output_path: Path,
    *,
    parser_name: str,
    account_number: str,
    inflow_description: str,
    outflow_description: str,
) -> None:
    inflow = Transaction(
        transaction_date=date(2025, 1, 2),
        description=inflow_description,
        credit=Decimal("250000.00"),
        debit=Decimal("0"),
        balance=Decimal("500000.00"),
    )
    outflow = Transaction(
        transaction_date=date(2025, 1, 3),
        description=outflow_description,
        credit=Decimal("0"),
        debit=Decimal("100000.00"),
        balance=Decimal("400000.00"),
    )
    inflow_classified = ClassifiedTransaction(transaction=inflow, classification="Sales", confidence=0.92)
    outflow_classified = ClassifiedTransaction(transaction=outflow, classification="Goods", confidence=0.9)
    analysis = StatementAnalysis(
        all_transactions=[inflow, outflow],
        classified_transactions=[inflow_classified, outflow_classified],
        inflows=[inflow_classified],
        outflows=[outflow_classified],
        parser_name=parser_name,
        metadata=StatementMetadata(
            account_name="TEST ACCOUNT",
            account_number=account_number,
            currency="NGN",
            opening_balance=Decimal("250000.00"),
            total_credit=Decimal("250000.00"),
            total_debit=Decimal("100000.00"),
            closing_balance=Decimal("400000.00"),
            period_start=date(2025, 1, 1),
            period_end=date(2025, 1, 31),
        ),
    )
    ExcelExporter().export(analysis, output_path)


if __name__ == "__main__":
    unittest.main()
