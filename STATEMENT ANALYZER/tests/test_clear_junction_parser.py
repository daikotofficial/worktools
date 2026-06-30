from __future__ import annotations

import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from statement_analyzer.models import TransactionDirection
from statement_analyzer.parsers.clear_junction import ClearJunctionStatementParser
from statement_analyzer.service import StatementAnalysisService


PROJECT_ROOT = Path(__file__).resolve().parents[1]


class ClearJunctionParserTests(unittest.TestCase):
    def test_signed_amount_statement_splits_inflows_outflows_and_fees(self) -> None:
        pdf_path = PROJECT_ROOT / "newstatementfile.pdf"
        parser = ClearJunctionStatementParser()

        self.assertTrue(parser.can_parse(pdf_path))
        transactions = parser.parse(pdf_path)
        transaction_rows = [
            transaction
            for transaction in transactions
            if transaction.direction in {TransactionDirection.INFLOW, TransactionDirection.OUTFLOW}
        ]

        self.assertEqual(len(transaction_rows), 563)
        self.assertEqual(parser.last_metadata.account_name, "SINAZ EXCHANGE BROKERS")
        self.assertEqual(parser.last_metadata.account_number, "GB48CLJU00997129900570")
        self.assertEqual(parser.last_metadata.currency, "EUR")
        self.assertEqual(parser.last_metadata.opening_balance, Decimal("0.02"))
        self.assertEqual(parser.last_metadata.closing_balance, Decimal("0.01"))
        self.assertEqual(parser.last_metadata.total_credit, Decimal("2834744.49"))
        self.assertEqual(parser.last_metadata.total_debit, Decimal("2829029.00"))

        first = transaction_rows[0]
        self.assertEqual(first.transaction_date, date(2025, 11, 1))
        self.assertEqual(first.credit, Decimal("12000.00"))
        self.assertEqual(first.debit, Decimal("0"))
        self.assertEqual(first.transaction_fee, Decimal("24.00"))
        self.assertIn("Debt refund", first.description)
        self.assertIn("ERICA VANESSA BRAS PINHEIRO", first.description)

        second = transaction_rows[1]
        self.assertEqual(second.credit, Decimal("0"))
        self.assertEqual(second.debit, Decimal("11975.00"))
        self.assertEqual(second.transaction_fee, Decimal("0.50"))
        self.assertIn("CEX.IO EU VASP, UAB", second.description)

        self.assertEqual(transaction_rows[-1].balance, Decimal("0.01"))

    def test_clear_junction_export_uses_simple_columns(self) -> None:
        pdf_path = PROJECT_ROOT / "newstatementfile.pdf"
        service = StatementAnalysisService()

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "clear_junction.xlsx"
            result = service.analyze(pdf_path, output_path)
            workbook = load_workbook(output_path, read_only=True, data_only=False)

            self.assertEqual(result.summary.parser_name, "clear-junction")
            self.assertEqual(result.summary.available_check_count, 4)
            self.assertEqual(result.summary.matched_check_count, 4)
            self.assertEqual(result.summary.review_rows, [])
            self.assertEqual(workbook.sheetnames, ["Transactions", "Inflows", "Outflows"])
            headers = [cell.value for cell in next(workbook["Transactions"].iter_rows(max_row=1))]

        self.assertEqual(headers, ["DATE", "DESCRIPTION", "INFLOW", "OUTFLOW", "TRANSACTION FEE"])


if __name__ == "__main__":
    unittest.main()
