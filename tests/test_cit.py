from __future__ import annotations

import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection

from statement_analyzer.cit import (
    CitSourceAmount,
    CitTargetCell,
    discover_template_targets,
    extract_accounting_year,
    parse_money,
    populate_cit_workbook,
    source_amount_from_row,
    source_amount_from_text_line,
    propose_mappings,
)


class CitTemplateTests(unittest.TestCase):
    def test_parse_money_handles_common_afs_formats(self) -> None:
        self.assertEqual(parse_money("1,234,567"), Decimal("1234567"))
        self.assertEqual(parse_money("(26,356,077)"), Decimal("-26356077"))
        self.assertIsNone(parse_money("NIL"))
        self.assertIsNone(parse_money("-"))

    def test_discovers_and_writes_only_unlocked_template_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "template.xlsx"
            output_path = Path(temp_dir) / "output.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "INCOME_STATEMENT"
            worksheet["C27"] = "Others"
            worksheet["D27"] = None
            worksheet["D27"].protection = Protection(locked=False)
            worksheet.protection.sheet = True
            workbook.save(template_path)

            targets = discover_template_targets(template_path)

            self.assertEqual([target.key for target in targets], ["INCOME_STATEMENT!D27"])

            populate_cit_workbook(template_path, output_path, {"INCOME_STATEMENT!D27": Decimal("5000")})
            generated = load_workbook(output_path, data_only=True)
            self.assertEqual(generated["INCOME_STATEMENT"]["D27"].value, 5000)

            with self.assertRaisesRegex(ValueError, "not an editable"):
                populate_cit_workbook(template_path, output_path, {"INCOME_STATEMENT!D28": Decimal("5000")})

    def test_proposes_expected_note_mappings(self) -> None:
        targets = [
            CitTargetCell("USER_INPUT", "D7", "ACCOUNTING YEAR", "user_input"),
            CitTargetCell("INCOME_STATEMENT", "D27", "Others", "revenue"),
            CitTargetCell("ADMINISTRATIVE_EXPENSES", "C21", "Audit fees", "administrative_expenses"),
            CitTargetCell("ADMINISTRATIVE_EXPENSES", "C4", "Salary and Wages", "administrative_expenses"),
        ]
        sources = [
            CitSourceAmount("Accounting year", Decimal("2025"), "user_input", 0, ("Accounting year", "2025")),
            CitSourceAmount("Total", Decimal("28182720"), "revenue_note", 12, ("", "Total", "", "28,182,720")),
            CitSourceAmount(
                "Audit & Tax Consultancy Fee",
                Decimal("250000"),
                "administrative_expenses_note",
                13,
                ("", "Audit & Tax Consultancy Fee", "", "250,000"),
            ),
            CitSourceAmount(
                "Personnel cost",
                Decimal("1250000"),
                "administrative_expenses_note",
                13,
                ("", "Personnel cost", "", "1,250,000"),
            ),
        ]

        proposals = propose_mappings(sources, targets)

        self.assertEqual(
            [proposal.target.key for proposal in proposals],
            ["USER_INPUT!D7", "INCOME_STATEMENT!D27", "ADMINISTRATIVE_EXPENSES!C21", "ADMINISTRATIVE_EXPENSES!C4"],
        )

    def test_text_extraction_repairs_spaced_current_year_amounts(self) -> None:
        source = source_amount_from_text_line(
            "WHT Credit 2 1 0,125,737 8 ,605,638",
            section="financial_position",
            page=7,
        )

        self.assertIsNotNone(source)
        self.assertEqual(source.label, "WHT Credit")
        self.assertEqual(source.amount, Decimal("10125737"))

    def test_text_extraction_skips_note_only_numbers(self) -> None:
        source = source_amount_from_text_line("Tax Provision 8", section="payables_note", page=7)

        self.assertIsNone(source)

    def test_row_extraction_uses_current_year_column(self) -> None:
        source = source_amount_from_row(
            ("Trade Payable", "7", "2,000,000", "700,000"),
            section="payables_note",
            page=7,
        )

        self.assertIsNotNone(source)
        self.assertEqual(source.label, "Trade Payable")
        self.assertEqual(source.amount, Decimal("2000000"))

    def test_cost_proposals_are_normalized_to_positive_amounts(self) -> None:
        targets = [CitTargetCell("INCOME_STATEMENT", "D84", "Other Direct costs", "cost_of_sales")]
        sources = [CitSourceAmount("Cost of Sales", Decimal("-1205834508"), "cost_note", 8, ("Cost of Sales",))]

        proposals = propose_mappings(sources, targets)

        self.assertEqual(proposals[0].source.amount, Decimal("1205834508"))

    def test_extract_accounting_year_prefers_template_or_file_name(self) -> None:
        self.assertEqual(
            extract_accounting_year(
                template_filename="2025_2521513338189_CIT.xlsx",
                afs_filename="AFS DAKADS.pdf",
                afs_path=Path("missing.pdf"),
            ),
            2025,
        )


if __name__ == "__main__":
    unittest.main()
